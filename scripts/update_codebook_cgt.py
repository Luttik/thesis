"""Update code-book.xlsx interpretive labels for CGT realignment."""

from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
XLSX = ROOT / "code-book.xlsx"

# row number (1-based in sheet) -> new interpretive / aggregated coding value
INTERPRETIVE_UPDATES: dict[int, str] = {
    2: "Experimenting",
    9: "Mobilising infrastructure & access",
    10: "Mobilising infrastructure & access",
    11: "Mobilising infrastructure & access",
    12: "Building specification capability",
    13: "",  # external experts — not internal resource
    14: "Enabling organizational conditions",
    15: "Navigating blocked conditions",
    16: "Navigating blocked conditions",
}

AGGREGATED_UPDATES: dict[int, str] = {
    2: "Enabling behavior",
    5: "Enabling behavior",
    9: "Mobilising internal conditions",
    10: "Mobilising internal conditions",
    11: "Mobilising internal conditions",
    12: "Mobilising internal conditions",
    14: "Mobilising internal conditions",
    15: "Experiencing obstacles",
    16: "Experiencing obstacles",
}


def main() -> None:
    wb = load_workbook(XLSX)
    ws = wb["Codes"]
    for row, val in INTERPRETIVE_UPDATES.items():
        ws.cell(row=row, column=2, value=val)
    for row, val in AGGREGATED_UPDATES.items():
        ws.cell(row=row, column=3, value=val)
    wb.save(XLSX)
    print(f"Updated {XLSX}")


if __name__ == "__main__":
    main()
