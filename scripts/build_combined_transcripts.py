"""Build a single combined transcripts document.

The combined document is **anonymised**: interviews are ordered by the
"Interviewee number" assigned in ``interviewees.xlsx`` (Sheet1), and each
section is titled ``Interviewee N`` instead of using the participant's
real name.

Run with:
    poetry run python scripts/build_combined_transcripts.py
"""

from __future__ import annotations

import re
import subprocess
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parent.parent
TRANSCRIPTS_DIR = ROOT / "transcripts" / "individuals"
OUTPUT_DIR = ROOT / "output"
INTERVIEWEES_XLSX = ROOT / "interviewees.xlsx"
COMBINED_MD = OUTPUT_DIR / "All Transcripts - Daan Luttik - MBA.md"
COMBINED_DOCX = OUTPUT_DIR / "All Transcripts - Daan Luttik - MBA.docx"

# Real name -> transcript markdown filename. The real name MUST match the
# value in ``interviewees.xlsx`` -> Sheet1 -> column B (case-insensitive).
TRANSCRIPT_FILES: dict[str, str] = {
    "Andreea Bulisache": "Thesis interview Andreea Bulisache.md",
    "Arjan Dijk": "Thesis interview Arjan Dijk.md",
    "Berfun Goodwin": "Thesis transcript Berfun Goodwin.md",
    "Dennis Goedegebuure": "Thesis Transcript Dennis Goedegebuure.md",
    "Erik Hilhorst": "Thesis Transcript Erik Hilhorst.md",
    "Erica": "Thesis interview Erica.md",
    "Georgio Mosis": "Thesis transcript Georgio Mosis.md",
    "Jon Stephan": "Thesis transcript Jon Stephan.md",
    "Lauren Stokowski": "Thesis interview Lauren Stokowski.md",
    "Maarten Mantjes": "Thesis transcript Maarten Mantjes.en.md",
    "Rolf Mulder": "Thesis transcript Rolf Mulder.md",
    "Scott Brinker": "Thesis Transcript Scott Brinker.md",
    "Tim Wiegel": "Thesis interview Tim Wiegel.md",
    "Floris Reguoin": "Thesis interview Floris Reguoin.md",
    "Sylvia Vroklage": "Thesis interview Sylvia Vroklage.md",
    # #16 is anonymous in ``interviewees.xlsx`` (e.g. "anonimous"/"(anonymous)").
    # Any real-name value that contains "anonym" is mapped to the Interview 16
    # markdown file below.
    "Karin Boon": "Thesis interview Karin Boon.md",
}
ANONYMOUS_TRANSCRIPT_FILE = "Thesis interview 16.md"


def _load_interviewee_numbers() -> list[tuple[int, str]]:
    """Return [(number, real_name), ...] from ``interviewees.xlsx`` -> Sheet1."""
    wb = load_workbook(INTERVIEWEES_XLSX, read_only=True, data_only=True)
    sheet = wb["Sheet1"]
    rows: list[tuple[int, str]] = []
    for raw_num, raw_name in sheet.iter_rows(min_row=2, max_col=2, values_only=True):
        if raw_num is None:
            continue
        try:
            number = int(raw_num)
        except (TypeError, ValueError):
            continue
        name = (raw_name or "").strip()
        rows.append((number, name))
    rows.sort(key=lambda r: r[0])
    return rows


def _resolve_transcript_file(number: int, name: str) -> str:
    if "anonym" in name.lower() or "anonim" in name.lower() or not name:
        return ANONYMOUS_TRANSCRIPT_FILE
    if name in TRANSCRIPT_FILES:
        return TRANSCRIPT_FILES[name]
    # Tolerant lookup: case-insensitive match against keys.
    lowered = {k.lower(): v for k, v in TRANSCRIPT_FILES.items()}
    if name.lower() in lowered:
        return lowered[name.lower()]
    raise KeyError(
        f"No transcript file mapping for interviewee #{number} ('{name}'). "
        f"Add it to TRANSCRIPT_FILES in {Path(__file__).name}."
    )


def _build_transcript_plan() -> list[tuple[int, str]]:
    """Return [(interviewee_number, transcript_filename), ...]."""
    plan: list[tuple[int, str]] = []
    for number, name in _load_interviewee_numbers():
        filename = _resolve_transcript_file(number, name)
        plan.append((number, filename))
    return plan

# Raw OpenXML page break – pandoc passes through raw blocks in `openxml` format.
PAGE_BREAK = (
    '\n```{=openxml}\n'
    '<w:p><w:r><w:br w:type="page"/></w:r></w:p>\n'
    "```\n\n"
)

# Match ATX headings (lines starting with one or more `#` followed by a space).
# We only demote headings that sit at column 0 and are not inside fenced code blocks.
_HEADING_RE = re.compile(r"^(#{1,5})\s")


def _demote_headings(markdown: str) -> str:
    """Add one '#' to every ATX heading outside fenced code blocks.

    This keeps the per-transcript H1 (added by this script) as the only
    top-level heading; any pre-existing H1/H2 inside a transcript becomes
    a subordinate heading.
    """
    out_lines: list[str] = []
    in_fence = False
    fence_marker: str | None = None
    for line in markdown.splitlines():
        stripped = line.lstrip()
        if stripped.startswith("```") or stripped.startswith("~~~"):
            marker = stripped[:3]
            if not in_fence:
                in_fence = True
                fence_marker = marker
            elif marker == fence_marker:
                in_fence = False
                fence_marker = None
            out_lines.append(line)
            continue

        if not in_fence:
            m = _HEADING_RE.match(line)
            if m:
                line = "#" + line
        out_lines.append(line)
    return "\n".join(out_lines)


def build_combined_markdown() -> Path:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    plan = _build_transcript_plan()

    parts: list[str] = []
    # Title page
    parts.append("---\n")
    parts.append('title: "Interview Transcripts"\n')
    parts.append('subtitle: "Thesis — Daan Luttik (MBA)"\n')
    parts.append("toc: true\n")
    parts.append("toc-depth: 1\n")
    parts.append("---\n\n")

    for position, (number, filename) in enumerate(plan):
        md_path = TRANSCRIPTS_DIR / filename
        if not md_path.is_file():
            raise FileNotFoundError(f"Missing transcript markdown: {md_path}")

        if position > 0:
            parts.append(PAGE_BREAK)

        parts.append(f"# Interviewee {number}\n\n")

        body = md_path.read_text(encoding="utf-8").strip()
        body = _demote_headings(body)
        parts.append(body)
        parts.append("\n\n")

    COMBINED_MD.write_text("".join(parts), encoding="utf-8")
    return COMBINED_MD


def render_docx(md_path: Path) -> Path:
    cmd = [
        "pandoc",
        str(md_path),
        "-f",
        "markdown",
        "-t",
        "docx",
        "--toc",
        "--toc-depth=1",
        "-o",
        str(COMBINED_DOCX),
    ]
    print("Running:", " ".join(cmd))
    subprocess.run(cmd, check=True)
    return COMBINED_DOCX


def main() -> None:
    md_path = build_combined_markdown()
    docx_path = render_docx(md_path)
    md_size_kb = md_path.stat().st_size / 1024
    docx_size_kb = docx_path.stat().st_size / 1024
    print(f"Wrote {md_path} ({md_size_kb:,.1f} KB)")
    print(f"Wrote {docx_path} ({docx_size_kb:,.1f} KB)")


if __name__ == "__main__":
    main()
